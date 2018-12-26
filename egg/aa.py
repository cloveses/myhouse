from openpyxl import load_workbook

datas = None
filename = 'a.xlsx'

def load():
    datas = []
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            # print()
            row_data.append(cell.value)
        datas.append(row_data)
    return datas

def get_ws():
    wb = load_workbook(filename)
    ws = wb.active
    return ws,wb

datas = load()

def add(row_data):
    ws,wb = get_ws()
    ws.append(row_data)
    wb.save(filename)
    datas = load()

def edit(seq,colname,data):
    ws,wb = get_ws()
    for row in ws.iter_rows():
        if row[0].value == seq:
            row[datas[0].index(colname)].value = data
            break
    wb.save(filename)
    datas = load()

def query(colname,data):
    datas = load()
    index = datas[0].index(colname)
    for data in datas:
        if data[index] == data:
            print(data)

def del_data(seq):
    ws,wb = get_ws()
    for row in ws.iter_rows():
        if row[0].value == seq:
            row[-1].value = '已售空'
    wb.save(filename)
    datas = load()

def main():
    print('commands:add,edi,que,del')
    while True:
        op = input('Pleas input a command:')
        if op == 'add':
            row_data = input('Please input data:')
            row_data = row_data.split(',')
            row_data[0] = int(row_data[0])
            row_data[3] = int(row_data[3])
        elif op == 'edi':
            edit_data = input('data:')
            edit_data = edit_data.split(',')
            edit(*edit_data)
        elif op == 'que':
            datass = input('data:')
            datass = datass.split(',')
            query(*datass)
        elif op == 'del':
            datass = input('Please input seq:')
            datass = int(datass)
            del_data(datass)
        elif op == 'exit':
            break
        else:
            continue

if __name__ == '__main__':
    # main()
    add((2,'aaa','bbb',30,'ccc'))
